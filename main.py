import openpyxl as xl
import pathlib
from openpyxl.chart import BarChart,Reference


#input file name with  extension
def process_work_book(filename):
    try:
        current_path = pathlib.Path().resolve()
        work_book = xl.load_workbook(f"{current_path}\ExcelFile\{filename}")
        sheet = work_book['Sheet1']



        #Loop through rows
        #include only rows with actual data and not headings
        for row in range(2,sheet.max_row + 1):
            #10% reduction of price in colum 3
            corrected_price = sheet.cell(row,3).value * 0.9
            corrected_price_cell = sheet.cell(row,4)
            corrected_price_cell.value = corrected_price

        #Refrence class used to select a range of values for chart
        values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)

        #Bar Chart
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart,'a6')
        #save files in saved file folder
        work_book.save(f'{current_path}\ExcelFile\savedFiles\{filename}')
        print("File saved...")
    except FileNotFoundError:
        print("No Such File Exist in directory..")
    except Exception as e:
        print(e)

#Main Body
print("Note: Place the File you wanted to Process in ExcelFile Folder. The Saved Result will be in SavedFiles Folder with the same name")
process_work_book(input("Enter Filename (include Extension): "))